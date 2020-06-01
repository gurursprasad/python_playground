import openpyxl
wb = openpyxl.load_workbook('C:\\Users\\gurur\\PycharmProjects\\Learn_Python\\ExcelData\\TestLeaf.xlsx')
# print(wb.sheetnames)
# Find active sheet name
# print(wb.active)
# Activate control to another sheet
# wb['Sheet1'].active.title

# Create object for sheet
sh = wb['Personal_Details']
# print(sh.title)

# Fetch Row and Column Count
row = sh.max_row
column = sh.max_column

# Way1
# print(sh['B2'].value)
# print(sh['D2'].value)
for eachrow in sh['A2':'D5']:
    for cells in eachrow:
        print(cells.value)
print('*'*30)

# Way2
for row in range(2,row+1):
    for cells in range(1,column+1):
        val = sh.cell(row,cells)
        print(val.value)
print('*'*30)